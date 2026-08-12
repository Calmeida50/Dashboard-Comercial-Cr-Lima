#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
gerar_base_clientes.py — planilha de conferencia da base de clientes.

Lista todos os clientes com venda em 2026 ou 2025, com o vendedor que atende
e as empresas em que compram, deixando uma coluna CANAL em branco para o
Cristiano classificar (alimentar / farma / indireto).

Depois de preenchida, a coluna CANAL vira a base da analise por canal.

Saida: _saida/Base_Clientes_Canal.xlsx
"""
import os, re, json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

PROJ = os.path.dirname(os.path.abspath(__file__))
MESES = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago",
         "Set", "Out", "Nov", "Dez"]


def carregar():
    h = open(os.path.join(PROJ, "index.html"), encoding="utf-8").read()
    m = re.search(r"const\s+DADOS_EMBEDDED\s*=\s*", h)
    i = m.end(); d = 0; j = i; ins = False; esc = False
    while j < len(h):
        c = h[j]
        if esc: esc = False
        elif c == "\\": esc = True
        elif c == '"': ins = not ins
        elif not ins:
            if c == "{": d += 1
            elif c == "}":
                d -= 1
                if d == 0: break
        j += 1
    return json.loads(h[i:j + 1])


def main():
    D = carregar()
    cd = D["clientes_detalhado"]
    gr = D["empresas"]["GERAL"]["real"]
    ate = max((i for i in range(12) if gr[i] > 0), default=6) + 1

    # consolida por cliente: soma as empresas, guarda vendedor e empresas
    reg = {}
    for emp in cd:
        for vend, lista in (cd[emp] or {}).items():
            for c in lista:
                nome = (c.get("nome") or "").strip()
                if not nome:
                    continue
                v26 = sum((c.get("meses") or [0] * 12)[:ate])
                v25 = sum((c.get("meses25") or [0] * 12)[:ate])
                if v26 <= 0 and v25 <= 0:
                    continue
                a = reg.setdefault(nome, {
                    "nome": nome, "vendedores": set(), "empresas": set(),
                    "v26": 0.0, "v25": 0.0, "meses": [0.0] * 12})
                a["vendedores"].add(vend)
                a["empresas"].add(emp)
                a["v26"] += v26
                a["v25"] += v25
                for k in range(12):
                    a["meses"][k] += (c.get("meses") or [0] * 12)[k]

    linhas = []
    for nome, a in reg.items():
        pos = sum(1 for k in range(ate) if a["meses"][k] > 0)
        linhas.append({
            "CLIENTE": nome,
            "VENDEDOR": " / ".join(sorted(a["vendedores"])),
            "CANAL": "",
            "EMPRESAS": " / ".join(sorted(a["empresas"])),
            "QTD EMPRESAS": len(a["empresas"]),
            "VENDA 2026": round(a["v26"], 2),
            "VENDA 2025": round(a["v25"], 2),
            "VAR %": (a["v26"] / a["v25"] - 1) if a["v25"] > 0 else None,
            "MESES C/ COMPRA": pos,
            "SITUACAO": ("Ativo" if a["v26"] > 0 else "Só comprou em 2025"),
        })
    df = pd.DataFrame(linhas).sort_values("VENDA 2026", ascending=False)
    return df, ate


def escrever(df, ate):
    wb = Workbook()
    ws = wb.active
    ws.title = "Base de Clientes"

    AZ = "1E293B"
    fina = Side(style="thin", color="D0D7DE")
    grade = Border(left=fina, right=fina, top=fina, bottom=fina)

    ws["A1"] = "BASE DE CLIENTES — conferência e classificação por canal"
    ws["A1"].font = Font(bold=True, size=13, color="1E293B")
    ws["A2"] = ("Preencha a coluna CANAL: alimentar, farma ou indireto "
                "(distribuidor/atacado). Clientes com venda em 2026 ou 2025.")
    ws["A2"].font = Font(size=10, italic=True, color="64748B")

    cab = list(df.columns)
    for c, nome in enumerate(cab, 1):
        cel = ws.cell(row=4, column=c, value=nome)
        cel.font = Font(bold=True, color="FFFFFF", size=10)
        cel.fill = PatternFill("solid", fgColor=AZ)
        cel.alignment = Alignment(horizontal="center", vertical="center",
                                  wrap_text=True)
        cel.border = grade

    for i, (_, r) in enumerate(df.iterrows(), start=5):
        for c, nome in enumerate(cab, 1):
            v = r[nome]
            if pd.isna(v):
                v = None
            cel = ws.cell(row=i, column=c, value=v)
            cel.border = grade
            cel.font = Font(size=10)
            if nome in ("VENDA 2026", "VENDA 2025"):
                cel.number_format = 'R$ #,##0.00'
            if nome == "VAR %":
                cel.number_format = "0.0%"
                if isinstance(v, float):
                    cel.font = Font(size=10, bold=True,
                                    color="16A34A" if v >= 0 else "DC2626")
            if nome in ("QTD EMPRESAS", "MESES C/ COMPRA"):
                cel.alignment = Alignment(horizontal="center")
            if nome == "CANAL":
                cel.fill = PatternFill("solid", fgColor="FFF7CC")
            if nome == "SITUACAO" and v == "Só comprou em 2025":
                cel.font = Font(size=10, color="B45309", bold=True)

    fim = 4 + len(df)
    # lista suspensa no CANAL, para padronizar a digitacao
    dv = DataValidation(type="list",
                        formula1='"alimentar,farma,indireto"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    col_canal = cab.index("CANAL") + 1
    dv.add("%s5:%s%d" % (get_column_letter(col_canal),
                         get_column_letter(col_canal), fim))

    larg = [46, 20, 14, 34, 9, 16, 16, 10, 10, 20]
    for c, w in enumerate(larg[:len(cab)], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:%s%d" % (get_column_letter(len(cab)), fim)

    # aba de resumo, que se preenche sozinha conforme o CANAL for classificado
    ws2 = wb.create_sheet("Resumo por Canal")
    ws2["A1"] = "Resumo por canal (preenche sozinho conforme a coluna CANAL)"
    ws2["A1"].font = Font(bold=True, size=12)
    for c, t in enumerate(["CANAL", "CLIENTES", "VENDA 2026", "VENDA 2025"], 1):
        cel = ws2.cell(row=3, column=c, value=t)
        cel.font = Font(bold=True, color="FFFFFF", size=10)
        cel.fill = PatternFill("solid", fgColor=AZ)
        cel.border = grade
    for i, canal in enumerate(["alimentar", "farma", "indireto", "(em branco)"], start=4):
        ws2.cell(row=i, column=1, value=canal).border = grade
        alvo = '"%s"' % canal if canal != "(em branco)" else '""'
        ws2.cell(row=i, column=2,
                 value='=COUNTIF(\'Base de Clientes\'!C5:C%d,%s)' % (fim, alvo)).border = grade
        c3 = ws2.cell(row=i, column=3,
                      value='=SUMIF(\'Base de Clientes\'!C5:C%d,%s,\'Base de Clientes\'!F5:F%d)' % (fim, alvo, fim))
        c3.number_format = 'R$ #,##0.00'; c3.border = grade
        c4 = ws2.cell(row=i, column=4,
                      value='=SUMIF(\'Base de Clientes\'!C5:C%d,%s,\'Base de Clientes\'!G5:G%d)' % (fim, alvo, fim))
        c4.number_format = 'R$ #,##0.00'; c4.border = grade
    for c, w in enumerate([16, 12, 18, 18], 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    saida = os.path.join(PROJ, "_saida")
    os.makedirs(saida, exist_ok=True)
    caminho = os.path.join(saida, "Base_Clientes_Canal.xlsx")
    wb.save(caminho)
    return caminho


if __name__ == "__main__":
    df, ate = main()
    p = escrever(df, ate)
    print("clientes: %d" % len(df))
    print("  ativos em 2026      : %d" % (df.SITUACAO == "Ativo").sum())
    print("  so compraram em 2025: %d" % (df.SITUACAO != "Ativo").sum())
    print("\nvendedores:")
    for v, n in df.VENDEDOR.value_counts().head(15).items():
        print("   %-24s %d" % (v[:24], n))
    print("\ngerado: %s" % p)
