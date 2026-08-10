#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
painel_renner.py — gera o painel gerencial da Renner em Excel.

ABAS:
  1. Resumo Gerencial   indicadores do topo + serie mensal reconstruida
  2. Analise de Vendas  por SKU: semana, mes e ano vs mesmo periodo do ano
                        anterior, em valor e em unidades
  3. Estoque por Loja   lojas na vertical, SKUs na horizontal (so as 80)
  4. Estoque E-commerce a loja 574, com o mix completo (20 itens)

REGRAS:
  - faturamento conta as 102 lojas (80 oficiais + 22 que receberam por engano)
  - estoque e ruptura olham SO as 80 oficiais
  - os 4 itens que saem de loja (Infusao Botanica, Folia, Expedicao, Elixir
    1870) ficam no FIM da analise, destacados, e FORA do calculo de ruptura
  - a loja 88 nunca recebeu produto: aparece marcada, no radar

Uso:  python3 painel_renner.py
Saida: _saida/Painel_Renner_Semana_NN.xlsx
"""
import os, sys, datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import renner as R

PROJ = os.path.dirname(os.path.abspath(__file__))

AZUL = "1E293B"; CINZA_H = "F1F5F9"; VERDE = "16A34A"; VERM = "DC2626"
AMAR = "FEF3C7"; LARANJA = "B45309"; BORDA = "E2E8F0"
fina = Side(style="thin", color=BORDA)
GRADE = Border(left=fina, right=fina, top=fina, bottom=fina)


def st_cab(tam=10):
    return {"font": Font(bold=True, color="FFFFFF", size=tam),
            "fill": PatternFill("solid", fgColor=AZUL),
            "alignment": Alignment(horizontal="center", vertical="center",
                                   wrap_text=True),
            "border": GRADE}


def aplica(cel, d):
    for k, v in d.items():
        setattr(cel, k, v)


def titulo(ws, texto, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=texto)
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(row=2, column=2 - 1, value=sub)
    c.font = Font(size=10, italic=True, color="64748B")


def aba_resumo(wb, sem, d, meses):
    ws = wb.create_sheet("Resumo Gerencial")
    of, ex, ec = R.grupos(d)
    ofp = of[of["cod"].isin(R.PERFUME_LOJA)]
    exp_ = ex[ex["cod"].isin(R.PERFUME_LOJA)]

    titulo(ws, "PAINEL GERENCIAL — RENNER / GRANADO",
           "Perfumes · semana %s · lojas físicas e e-commerce" % sem, 8)

    # --- cartoes
    ytd_of = ofp["vl_ytd"].sum(); ytd_ex = exp_["vl_ytd"].sum()
    ytd_ec = ec["vl_ytd"].sum()
    ytd_aa = ofp["vl_ytd_aa"].sum() + exp_["vl_ytd_aa"].sum() + ec["vl_ytd_aa"].sum()
    ytd_tot = ytd_of + ytd_ex + ytd_ec
    sem_tot = ofp["vl_sem"].sum() + exp_["vl_sem"].sum() + ec["vl_sem"].sum()
    sem_aa = ofp["vl_sem_aa"].sum() + exp_["vl_sem_aa"].sum() + ec["vl_sem_aa"].sum()
    est_of = ofp["estoque"].sum(); tr_of = ofp["transito"].sum()
    # ruptura: so oficiais, so itens que ficam em loja
    pares = len(ofp)
    zer = int((ofp["estoque"] <= 0).sum())
    rup = zer / pares * 100 if pares else 0

    cards = [
        ("VENDA SEMANA", sem_tot, "R$ #,##0.00",
         ("▲ " if sem_tot >= sem_aa else "▼ ") +
         ("%.1f%% vs mesma semana AA" % (abs(sem_tot / sem_aa - 1) * 100) if sem_aa else "sem base AA")),
        ("VENDA YTD", ytd_tot, "R$ #,##0.00",
         ("▲ " if ytd_tot >= ytd_aa else "▼ ") +
         ("%.1f%% vs AA" % (abs(ytd_tot / ytd_aa - 1) * 100) if ytd_aa else "sem base AA")),
        ("ESTOQUE — 80 LOJAS", est_of, "#,##0", "unidades"),
        ("EM TRÂNSITO", tr_of, "#,##0", "unidades"),
        ("RUPTURA — 80 LOJAS", rup / 100, "0.0%", "%d de %d pares loja-SKU" % (zer, pares)),
        ("E-COMMERCE YTD", ytd_ec, "R$ #,##0.00",
         "%.1f%% do total" % (ytd_ec / ytd_tot * 100 if ytd_tot else 0)),
    ]
    lin = 4
    for i, (rot, val, fmt, obs) in enumerate(cards):
        col = 1 + (i % 3) * 3
        l = lin + (i // 3) * 4
        ws.merge_cells(start_row=l, start_column=col, end_row=l, end_column=col + 2)
        c = ws.cell(row=l, column=col, value=rot)
        aplica(c, st_cab(9))
        ws.merge_cells(start_row=l + 1, start_column=col, end_row=l + 1, end_column=col + 2)
        c = ws.cell(row=l + 1, column=col, value=val)
        c.font = Font(bold=True, size=15)
        c.number_format = fmt
        c.alignment = Alignment(horizontal="center")
        c.border = GRADE
        ws.merge_cells(start_row=l + 2, start_column=col, end_row=l + 2, end_column=col + 2)
        c = ws.cell(row=l + 2, column=col, value=obs)
        c.font = Font(size=9, color="64748B")
        c.alignment = Alignment(horizontal="center")
        c.border = GRADE

    # --- serie mensal
    l = 13
    ws.cell(row=l, column=1, value="EVOLUÇÃO MENSAL (fechamento pela virada do MTD)").font = Font(bold=True, size=11)
    l += 1
    cab = ["Mês", "Semana fech.", "Venda R$", "Venda R$ AA", "Var. %",
           "Unidades", "Un. AA", "Var. un. %"]
    for j, t in enumerate(cab, 1):
        aplica(ws.cell(row=l, column=j, value=t), st_cab())
    for nome, s, v, vaa, u, uaa in meses:
        l += 1
        vals = [nome, s, v, vaa,
                (v / vaa - 1) if vaa else None, u, uaa,
                (u / uaa - 1) if uaa else None]
        for j, val in enumerate(vals, 1):
            c = ws.cell(row=l, column=j, value=val)
            c.border = GRADE
            c.font = Font(size=10)
            if j in (3, 4):
                c.number_format = "R$ #,##0.00"
            if j in (6, 7):
                c.number_format = "#,##0"
            if j in (5, 8):
                c.number_format = "0.0%"
                if val is not None:
                    c.font = Font(size=10, bold=True,
                                  color=VERDE if val >= 0 else VERM)
    for j, w in enumerate([18, 12, 15, 15, 10, 11, 10, 11], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    return ws


def aba_vendas(wb, sem, d):
    """por SKU: semana / mes / ano, valor e unidades, vs ano anterior.
    Os 4 itens que saem de loja vao no FIM, destacados."""
    ws = wb.create_sheet("Análise de Vendas")
    titulo(ws, "ANÁLISE DE VENDAS POR SKU",
           "Semana %s · todas as lojas + e-commerce · YTD em unidades não tem "
           "base do ano anterior no relatório" % sem, 15)

    g = d.groupby("cod").agg(
        item=("item", "first"),
        vl_sem=("vl_sem", "sum"), vl_sem_aa=("vl_sem_aa", "sum"),
        vl_mtd=("vl_mtd", "sum"), vl_mtd_aa=("vl_mtd_aa", "sum"),
        vl_ytd=("vl_ytd", "sum"), vl_ytd_aa=("vl_ytd_aa", "sum"),
        un_sem=("un_sem", "sum"), un_sem_aa=("un_sem_aa", "sum"),
        un_mtd=("un_mtd", "sum"), un_mtd_aa=("un_mtd_aa", "sum"),
        un_ytd=("un_ytd", "sum")).reset_index()
    g = g[g["cod"].isin(R.MIX_PERFUME)]
    g["nome"] = g["cod"].map(R.MIX_PERFUME)
    g["so_ecom"] = g["cod"].isin(R.PERFUME_SO_ECOM)
    g = pd.concat([g[~g.so_ecom].sort_values("vl_ytd", ascending=False),
                   g[g.so_ecom].sort_values("vl_ytd", ascending=False)])

    l = 4
    ws.merge_cells(start_row=l, start_column=3, end_row=l, end_column=6)
    ws.cell(row=l, column=3, value="VALOR R$")
    ws.merge_cells(start_row=l, start_column=7, end_row=l, end_column=10)
    ws.cell(row=l, column=7, value="UNIDADES")
    for j in (3, 7):
        aplica(ws.cell(row=l, column=j), st_cab())
    for j in range(4, 7):
        aplica(ws.cell(row=l, column=j), st_cab())
    for j in range(8, 11):
        aplica(ws.cell(row=l, column=j), st_cab())
    l += 1
    cab = ["Produto", "Código",
           "Semana", "Sem. AA", "MTD", "YTD",
           "Semana", "Sem. AA", "MTD", "YTD"]
    for j, t in enumerate(cab, 1):
        aplica(ws.cell(row=l, column=j, value=t), st_cab())

    for _, r in g.iterrows():
        l += 1
        vals = [r["nome"], r["cod"],
                r.vl_sem, r.vl_sem_aa, r.vl_mtd, r.vl_ytd,
                r.un_sem, r.un_sem_aa, r.un_mtd, r.un_ytd]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=l, column=j, value=v)
            c.border = GRADE
            c.font = Font(size=10)
            if 3 <= j <= 6:
                c.number_format = "R$ #,##0.00"
            if j >= 7:
                c.number_format = "#,##0"
            if r.so_ecom:
                c.fill = PatternFill("solid", fgColor=AMAR)
                if j == 1:
                    c.font = Font(size=10, bold=True, color=LARANJA)
    l += 1
    # linha de TOTAL
    tot = [g[c].sum() for c in ["vl_sem", "vl_sem_aa", "vl_mtd", "vl_ytd",
                                "un_sem", "un_sem_aa", "un_mtd", "un_ytd"]]
    vals = ["TOTAL", ""] + tot
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=l, column=j, value=v)
        c.border = GRADE
        c.font = Font(size=10, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=AZUL)
        if 3 <= j <= 6:
            c.number_format = "R$ #,##0.00"
        if j >= 7:
            c.number_format = "#,##0"

    l += 2
    ws.cell(row=l, column=1,
            value="Itens em amarelo saem das lojas físicas e ficam apenas no "
                  "e-commerce — não entram no cálculo de ruptura.")
    ws.cell(row=l, column=1).font = Font(size=9, italic=True, color=LARANJA)
    for j, w in enumerate([42, 12, 14, 13, 14, 15, 10, 10, 10, 10], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "C6"
    return ws


def aba_estoque_loja(wb, sem, d):
    """lojas na vertical, SKUs na horizontal — SO as 80 oficiais"""
    ws = wb.create_sheet("Estoque por Loja")
    of, _ex, _ec = R.grupos(d)
    ofp = of[of["cod"].isin(R.PERFUME_LOJA)]

    # SKUs ordenados do MAIOR para o menor por venda no ano, para os itens
    # mais relevantes ficarem a esquerda
    vend = ofp.groupby("cod")["vl_ytd"].sum().to_dict()
    skus = sorted(R.PERFUME_LOJA, key=lambda c: -vend.get(c, 0))
    piv = ofp.pivot_table(index=["loja_cod", "loja_nome"], columns="cod",
                          values="estoque", aggfunc="sum", fill_value=0)
    for c in skus:
        if c not in piv.columns:
            piv[c] = 0
    piv = piv[skus].reset_index()
    # lojas: as com MAIS ruptura primeiro — sao as que precisam de acao
    piv["_zer"] = piv[skus].apply(lambda r: int((r <= 0).sum()), axis=1)
    piv = piv.sort_values(["_zer", "loja_cod"], ascending=[False, True])
    piv = piv.drop(columns=["_zer"])

    titulo(ws, "ESTOQUE POR LOJA — 80 LOJAS OFICIAIS",
           "Semana %s · unidades por SKU · só itens que ficam em loja física" % sem,
           len(skus) + 5)
    l = 4
    ws.cell(row=l, column=1, value="Loja")
    ws.cell(row=l + 1, column=1, value="Código do SKU →")
    for j, c in enumerate(skus, 2):
        ws.cell(row=l, column=j, value=R.PERFUME_LOJA[c])
        ws.cell(row=l + 1, column=j, value=c)
    for j, t in enumerate(["Total", "SKUs em ruptura", "% Ruptura"],
                          len(skus) + 2):
        ws.cell(row=l, column=j, value=t)
        ws.cell(row=l + 1, column=j, value="")
    for j in range(1, len(skus) + 5):
        aplica(ws.cell(row=l, column=j), st_cab(9))
        aplica(ws.cell(row=l + 1, column=j), st_cab(8))
    ws.row_dimensions[l].height = 44

    lin = l + 1
    for _, r in piv.iterrows():
        lin += 1
        cod = int(r["loja_cod"])
        ws.cell(row=lin, column=1, value="%d - %s" % (cod, r["loja_nome"]))
        vals = [int(r[c]) for c in skus]
        for j, v in enumerate(vals, 2):
            c_ = ws.cell(row=lin, column=j, value=v)
            c_.number_format = "#,##0"
            if v <= 0:
                c_.fill = PatternFill("solid", fgColor="FEE2E2")
                c_.font = Font(size=10, bold=True, color=VERM)
        zer = sum(1 for v in vals if v <= 0)
        ws.cell(row=lin, column=len(skus) + 2, value=sum(vals)).number_format = "#,##0"
        ws.cell(row=lin, column=len(skus) + 3, value=zer)
        c_ = ws.cell(row=lin, column=len(skus) + 4, value=zer / len(skus))
        c_.number_format = "0.0%"
        c_.font = Font(size=10, bold=True,
                       color=VERM if zer / len(skus) > 0.3 else VERDE)
        for j in range(1, len(skus) + 5):
            ws.cell(row=lin, column=j).border = GRADE
            if j == 1:
                ws.cell(row=lin, column=j).font = Font(size=10)

    # loja que nunca recebeu produto
    for cod in sorted(R.LOJAS_SEM_HISTORICO):
        lin += 1
        ws.cell(row=lin, column=1,
                value="%d - (nunca recebeu produto — acompanhar)" % cod)
        for j in range(1, len(skus) + 5):
            c_ = ws.cell(row=lin, column=j)
            c_.fill = PatternFill("solid", fgColor=AMAR)
            c_.border = GRADE
            c_.font = Font(size=10, italic=True, color=LARANJA)

    # --- rodape por SKU: em quantas lojas tem, em quantas falta, % ruptura
    lin += 1
    n_lojas = len(piv)
    rotulos = [("LOJAS COM O PRODUTO", "com"),
               ("LOJAS EM RUPTURA", "rup"),
               ("% RUPTURA DO ITEM", "pct")]
    for rot, tipo in rotulos:
        lin += 1
        c = ws.cell(row=lin, column=1, value=rot)
        aplica(c, st_cab(9))
        c.alignment = Alignment(horizontal="right", vertical="center")
        for j, cod in enumerate(skus, 2):
            com = int((piv[cod] > 0).sum())
            rup = n_lojas - com
            if tipo == "com":
                v = com
            elif tipo == "rup":
                v = rup
            else:
                v = rup / n_lojas if n_lojas else 0
            cel = ws.cell(row=lin, column=j, value=v)
            cel.border = GRADE
            cel.alignment = Alignment(horizontal="center")
            if tipo == "pct":
                cel.number_format = "0.0%"
                cel.font = Font(size=10, bold=True,
                                color=VERM if v > 0.3 else VERDE)
            else:
                cel.number_format = "#,##0"
                cel.font = Font(size=10, bold=True,
                                color=VERM if tipo == "rup" and v > 0 else "1E293B")
        # total da linha
        cel = ws.cell(row=lin, column=len(skus) + 2)
        cel.border = GRADE
        if tipo == "pct":
            tot_rup = sum(n_lojas - int((piv[c_].sum() > 0)) for c_ in skus)
            zer_tot = sum(int((piv[c_] <= 0).sum()) for c_ in skus)
            cel.value = zer_tot / (n_lojas * len(skus))
            cel.number_format = "0.0%"
            cel.font = Font(size=10, bold=True, color=VERM)
        for j in range(len(skus) + 3, len(skus) + 5):
            ws.cell(row=lin, column=j).border = GRADE

    ws.column_dimensions["A"].width = 34
    for j in range(2, len(skus) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 11
    for j in range(len(skus) + 2, len(skus) + 5):
        ws.column_dimensions[get_column_letter(j)].width = 13
    ws.freeze_panes = ws.cell(row=l + 2, column=2)
    return ws


def aba_ecom(wb, sem, d):
    ws = wb.create_sheet("Estoque E-commerce")
    _of, _ex, ec = R.grupos(d)
    skus = list(R.PERFUME_LOJA) + list(R.PERFUME_SO_ECOM)
    titulo(ws, "E-COMMERCE — ESTOQUE E VENDA",
           "Semana %s · loja %d · mix completo (20 itens)" % (sem, R.LOJA_ECOM),
           7)
    l = 4
    cab = ["Produto", "Código", "Estoque", "Em trânsito",
           "Venda semana R$", "Venda MTD R$", "Venda YTD R$"]
    for j, t in enumerate(cab, 1):
        aplica(ws.cell(row=l, column=j, value=t), st_cab())
    g = ec.groupby("cod").agg(
        est=("estoque", "sum"), tr=("transito", "sum"),
        vs=("vl_sem", "sum"), vm=("vl_mtd", "sum"), vy=("vl_ytd", "sum")).to_dict("index")
    # ordena do MAIOR para o menor por venda YTD, mantendo os 4 de e-commerce
    # exclusivo no fim (destacados)
    em_loja = sorted([c for c in R.PERFUME_LOJA],
                     key=lambda c: -(g.get(c, {}).get("vy", 0)))
    so_ec = sorted([c for c in R.PERFUME_SO_ECOM],
                   key=lambda c: -(g.get(c, {}).get("vy", 0)))
    skus = em_loja + so_ec
    tot = {"est": 0, "tr": 0, "vs": 0, "vm": 0, "vy": 0}
    for c in skus:
        l += 1
        r = g.get(c, {})
        for k in tot:
            tot[k] += r.get(k, 0)
        so_ecom = c in R.PERFUME_SO_ECOM
        vals = [R.MIX_PERFUME[c], c, r.get("est", 0), r.get("tr", 0),
                r.get("vs", 0), r.get("vm", 0), r.get("vy", 0)]
        for j, v in enumerate(vals, 1):
            c_ = ws.cell(row=l, column=j, value=v)
            c_.border = GRADE
            c_.font = Font(size=10)
            if j in (3, 4):
                c_.number_format = "#,##0"
            if j >= 5:
                c_.number_format = "R$ #,##0.00"
            if so_ecom:
                c_.fill = PatternFill("solid", fgColor=AMAR)
                if j == 1:
                    c_.font = Font(size=10, bold=True, color=LARANJA)
    # TOTAL
    l += 1
    vals = ["TOTAL", "", tot["est"], tot["tr"], tot["vs"], tot["vm"], tot["vy"]]
    for j, v in enumerate(vals, 1):
        c_ = ws.cell(row=l, column=j, value=v)
        c_.border = GRADE
        c_.font = Font(size=10, bold=True, color="FFFFFF")
        c_.fill = PatternFill("solid", fgColor=AZUL)
        if j in (3, 4):
            c_.number_format = "#,##0"
        if j >= 5:
            c_.number_format = "R$ #,##0.00"
    for j, w in enumerate([42, 12, 11, 12, 16, 15, 16], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"
    return ws


def main():
    print("lendo as semanas da Renner...")
    sem, d = R.ultima_semana()
    if d is None:
        print("nenhuma semana com dados encontrada.")
        return 1
    meses = R.meses_fechados()
    print("semana mais recente: %s | meses reconstruidos: %d" % (sem, len(meses)))

    wb = Workbook()
    wb.remove(wb.active)
    aba_resumo(wb, sem, d, meses)
    aba_vendas(wb, sem, d)
    aba_estoque_loja(wb, sem, d)
    aba_ecom(wb, sem, d)

    saida = os.path.join(PROJ, "_saida")
    os.makedirs(saida, exist_ok=True)
    caminho = os.path.join(saida, "Painel_Renner_Semana_%s.xlsx" % sem)
    wb.save(caminho)
    print("gerado:", caminho)

    of, ex, ec = R.grupos(d)
    ofp = of[of["cod"].isin(R.PERFUME_LOJA)]
    print("\n  80 oficiais : %d lojas | estoque %s un | ruptura %.1f%%"
          % (ofp["loja_cod"].nunique(), "{:,.0f}".format(ofp["estoque"].sum()),
             (ofp["estoque"] <= 0).sum() / len(ofp) * 100))
    print("  extras      : %d lojas | venda YTD %s"
          % (ex[ex["cod"].isin(R.PERFUME_LOJA)]["loja_cod"].nunique(),
             "{:,.2f}".format(ex[ex["cod"].isin(R.PERFUME_LOJA)]["vl_ytd"].sum())))
    print("  e-commerce  : venda YTD %s" % "{:,.2f}".format(ec["vl_ytd"].sum()))
    return 0


if __name__ == "__main__":
    sys.exit(main())
